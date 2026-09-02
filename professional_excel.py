#!/usr/bin/env python
"""Create a polished, human-readable Excel workbook from experiment CSV data."""

import argparse
import csv
from collections import Counter
from pathlib import Path
import re
import sys


FRIENDLY_HEADERS = {
    "participant": "Participant ID",
    "session": "Session",
    "date": "Date & Time",
    "age": "Age",
    "gender": "Gender",
    "trial_number": "Trial Number",
    "load_condition": "Cognitive Load Level",
    "change_condition": "Change Condition",
    "changed_position": "Changed Square Position",
    "original_color_1": "Original Color 1",
    "original_color_2": "Original Color 2",
    "original_color_3": "Original Color 3",
    "original_color_4": "Original Color 4",
    "test_color_1": "Test Color 1",
    "test_color_2": "Test Color 2",
    "test_color_3": "Test Color 3",
    "test_color_4": "Test Color 4",
    "parity_digit_1": "Parity Digit 1",
    "parity_response_1": "Parity Response 1",
    "parity_correct_1": "Parity Result 1",
    "parity_rt_1": "Parity Reaction Time 1 (seconds)",
    "parity_digit_2": "Parity Digit 2",
    "parity_response_2": "Parity Response 2",
    "parity_correct_2": "Parity Result 2",
    "parity_rt_2": "Parity Reaction Time 2 (seconds)",
    "parity_digit_3": "Parity Digit 3",
    "parity_response_3": "Parity Response 3",
    "parity_correct_3": "Parity Result 3",
    "parity_rt_3": "Parity Reaction Time 3 (seconds)",
    "parity_digit_4": "Parity Digit 4",
    "parity_response_4": "Parity Response 4",
    "parity_correct_4": "Parity Result 4",
    "parity_rt_4": "Parity Reaction Time 4 (seconds)",
    "memory_response": "Memory Response",
    "memory_correct": "Memory Result",
    "memory_rt": "Memory Reaction Time (seconds)",
    "thought_probe_response": "Thought Category Number",
    "thought_probe_label": "Thought Category",
    "thought_probe_rt": "Thought Probe Reaction Time (seconds)",
    "trial_start_time": "Trial Start Time (seconds)",
    "trial_end_time": "Trial End Time (seconds)",
    "trial_duration": "Trial Duration (seconds)",
}

THOUGHT_LABELS = {
    "task": "Task",
    "task experience/performance": "Task Experience / Performance",
    "everyday things": "Everyday Things",
    "current state of being": "Current State of Being",
    "personal worries": "Personal Worries",
    "daydreams": "Daydreams",
    "external environment": "External Environment",
    "other": "Other",
}

NAVY = "17365D"
BLUE = "DCE6F1"
GOLD = "FFF2CC"
PURPLE = "E4DFEC"
GREEN = "E2F0D9"
RED = "FCE4D6"
LIGHT = "F7F9FC"
BORDER = "CBD5E1"
TEXT = "1F2937"


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _mean(values):
    values = [value for value in values if value is not None]
    return sum(values) / len(values) if values else None


def _friendly_date(value):
    match = re.match(r"^(\d{4}-\d{2}-\d{2})_(\d{2})h(\d{2})\.(\d{2})", value or "")
    if match:
        return f"{match.group(1)} {match.group(2)}:{match.group(3)}:{match.group(4)}"
    return (value or "").replace("_", " ")


def _friendly_value(header, value):
    value = "" if value is None else str(value).strip()
    upper = value.upper()
    if upper in {"NA", "N/A", "NONE", ""}:
        if header == "changed_position":
            return "No Change"
        if header.startswith("parity_"):
            return "Not Required"
        if header.startswith("memory_") or header.startswith("thought_probe_"):
            return "No Response"
        return "Not Recorded"

    if header == "date":
        return _friendly_date(value)
    if header == "load_condition":
        return f"Load {value}"
    if header == "change_condition":
        return "No Change" if value.lower() == "no_change" else "Change"
    if header == "gender":
        return value.replace("_", " ").title()
    if header.startswith("original_color_") or header.startswith("test_color_"):
        return value.title()
    if header.startswith("parity_response_"):
        return {"f": "Odd (F)", "j": "Even (J)"}.get(value.lower(), value.upper())
    if header == "memory_response":
        return {"s": "Same (S)", "d": "Different (D)"}.get(value.lower(), value.upper())
    if "correct" in header:
        return "Correct" if value in {"1", "1.0", "True", "true"} else "Incorrect"
    if header == "thought_probe_label":
        normalized = value.replace("_", " ").strip().lower()
        return THOUGHT_LABELS.get(normalized, value.replace("_", " "))

    if header in {"age", "trial_number", "changed_position", "thought_probe_response"} or header.startswith("parity_digit_"):
        number = _number(value)
        return int(number) if number is not None and number.is_integer() else value
    if header.endswith("_rt") or "_rt_" in header or header in {"trial_start_time", "trial_end_time", "trial_duration"}:
        number = _number(value)
        return number if number is not None else value
    return value


def _experiment_name(csv_path):
    name = csv_path.stem.lower()
    if "experiment1" in name:
        return "Experiment 1 — Simultaneous"
    if "experiment2" in name:
        return "Experiment 2 — Sequential"
    return "Working Memory Experiment"


def _load_csv(csv_path):
    with csv_path.open("r", newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        headers = reader.fieldnames or []
        rows = list(reader)
    if not headers:
        raise ValueError("The CSV file has no header row.")
    return headers, rows


def _summary_values(rows):
    memory_correct = [_number(row.get("memory_correct")) for row in rows]
    memory_rt = [_number(row.get("memory_rt")) for row in rows]
    parity_correct = []
    parity_rt = []
    for row in rows:
        for index in range(1, 5):
            parity_correct.append(_number(row.get(f"parity_correct_{index}")))
            parity_rt.append(_number(row.get(f"parity_rt_{index}")))
    thought_counts = Counter(
        row.get("thought_probe_label", "").strip()
        for row in rows
        if row.get("thought_probe_label", "").strip().upper() not in {"", "NA", "N/A"}
    )
    most_common_thought = thought_counts.most_common(1)[0][0] if thought_counts else "No Data"
    return {
        "trials": len(rows),
        "memory_accuracy": _mean(memory_correct),
        "memory_rt": _mean(memory_rt),
        "parity_accuracy": _mean(parity_correct),
        "most_common_thought": THOUGHT_LABELS.get(
            most_common_thought.replace("_", " ").strip().lower(),
            most_common_thought.replace("_", " "),
        ),
    }


def create_professional_workbook(csv_file, output_file=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("PsychoPy's Excel component (openpyxl) is not installed.") from exc

    csv_path = Path(csv_file).resolve()
    output_path = Path(output_file).resolve() if output_file else csv_path.with_suffix(".xlsx")
    headers, rows = _load_csv(csv_path)
    summary_values = _summary_values(rows)
    first = rows[0] if rows else {}

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    trials = wb.create_sheet("Trial Data")
    wb.properties.creator = "PsychoPy Working Memory Experiment"
    wb.properties.title = f"{_experiment_name(csv_path)} Results"
    wb.properties.subject = "Professional experiment results workbook"

    thin = Side(style="thin", color=BORDER)
    section_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    summary.sheet_view.showGridLines = False
    summary.sheet_properties.tabColor = NAVY
    summary.merge_cells("A1:H2")
    summary["A1"] = "WORKING MEMORY EXPERIMENT RESULTS"
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF", underline=None)
    summary["A1"].alignment = centered
    for row in summary["A1:H2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)

    metadata = [
        ("Participant ID", first.get("participant", "No Data")),
        ("Experiment", _experiment_name(csv_path)),
        ("Session", first.get("session", "No Data")),
        ("Date", _friendly_date(first.get("date", "")) or "No Data"),
        ("Participant Profile", f"Age {first.get('age', '—')} · {str(first.get('gender', '—')).title()}"),
    ]
    for row_index, (label, value) in enumerate(metadata, start=4):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        summary.cell(row_index, 1).font = Font(bold=True, color=NAVY, underline=None)
        for column in (1, 2):
            cell = summary.cell(row_index, column)
            cell.alignment = centered
            cell.border = section_border
            if column == 2:
                cell.font = Font(color=TEXT, underline=None)
                if label in {"Participant ID", "Session"}:
                    cell.number_format = "@"

    summary.merge_cells("D4:H4")
    summary["D4"] = "PERFORMANCE OVERVIEW"
    for row in summary["D4:H4"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)
    summary["D4"].font = Font(bold=True, color="FFFFFF", underline=None)
    summary["D4"].alignment = centered

    overview_headers = [
        "Trials Completed",
        "Memory Accuracy",
        "Average Memory Reaction Time",
        "Parity Accuracy",
        "Most Frequent Thought Category",
    ]
    overview_values = [
        summary_values["trials"],
        summary_values["memory_accuracy"] if summary_values["memory_accuracy"] is not None else "No Data",
        summary_values["memory_rt"] if summary_values["memory_rt"] is not None else "No Data",
        summary_values["parity_accuracy"] if summary_values["parity_accuracy"] is not None else "No Data",
        summary_values["most_common_thought"],
    ]
    for column_index, value in enumerate(overview_headers, start=4):
        cell = summary.cell(5, column_index, value)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True, color=NAVY, underline=None)
        cell.alignment = centered
        cell.border = section_border
    for column_index, value in enumerate(overview_values, start=4):
        cell = summary.cell(6, column_index, value)
        cell.font = Font(size=13, bold=True, color=TEXT, underline=None)
        cell.alignment = centered
        cell.border = section_border
    if isinstance(summary["E6"].value, (int, float)):
        summary["E6"].number_format = "0.0%"
    if isinstance(summary["F6"].value, (int, float)):
        summary["F6"].number_format = '0.000 "s"'
    if isinstance(summary["G6"].value, (int, float)):
        summary["G6"].number_format = "0.0%"

    summary.merge_cells("A11:H11")
    summary["A11"] = "COLOR GUIDE"
    for row in summary["A11:H11"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)
    summary["A11"].font = Font(bold=True, color="FFFFFF", underline=None)
    summary["A11"].alignment = centered
    guide = [
        ("Load 0", "No parity questions", BLUE, "Correct", "Correct response", GREEN),
        ("Load 2", "Two parity questions", GOLD, "Incorrect", "Incorrect or missed response", RED),
        ("Load 4", "Four parity questions", PURPLE, "No Response", "Participant did not answer", RED),
        ("Raw Data", "CSV preserved for analysis", "FFFFFF", "Not Required", "Question was not presented", LIGHT),
    ]
    for row_index, (left_label, left_text, left_fill, right_label, right_text, right_fill) in enumerate(guide, start=12):
        values = {1: left_label, 2: left_text, 5: right_label, 6: right_text}
        for column_index in range(1, 9):
            cell = summary.cell(row_index, column_index, values.get(column_index, ""))
            cell.fill = PatternFill("solid", fgColor=left_fill if column_index <= 4 else right_fill)
            cell.font = Font(color=TEXT, underline=None)
            cell.alignment = centered
            cell.border = section_border

    summary.merge_cells("A17:H17")
    summary["A17"] = "Generated automatically after the experiment closes. All reaction times are reported in seconds."
    summary["A17"].fill = PatternFill("solid", fgColor=LIGHT)
    summary["A17"].font = Font(italic=True, color="475569", underline=None)
    summary["A17"].alignment = centered
    for column in range(1, 9):
        summary.column_dimensions[get_column_letter(column)].width = 20
    summary.row_dimensions[1].height = 34
    summary.row_dimensions[2].height = 34
    for row in range(3, 18):
        summary.row_dimensions[row].height = 24
    summary.freeze_panes = "A4"
    summary.sheet_view.zoomScale = 90

    trials.sheet_view.showGridLines = False
    trials.sheet_properties.tabColor = NAVY
    friendly_headers = [FRIENDLY_HEADERS.get(header, header.replace("_", " ").title()) for header in headers]
    trials.append(friendly_headers)
    for row in rows:
        trials.append([_friendly_value(header, row.get(header)) for header in headers])

    header_fill = PatternFill("solid", fgColor=NAVY)
    for cell in trials[1]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF", underline=None)
        cell.alignment = centered
        cell.border = Border(bottom=Side(style="medium", color="FFFFFF"))
    trials.row_dimensions[1].height = 48
    for text_header in ("participant", "session"):
        if text_header in headers:
            text_column = headers.index(text_header) + 1
            for row_index in range(2, trials.max_row + 1):
                trials.cell(row_index, text_column).number_format = "@"

    row_fills = {"0": BLUE, "2": GOLD, "4": PURPLE}
    load_index = headers.index("load_condition") + 1 if "load_condition" in headers else None
    result_indexes = [index + 1 for index, header in enumerate(headers) if "correct" in header]
    rt_indexes = [
        index + 1
        for index, header in enumerate(headers)
        if header.endswith("_rt") or "_rt_" in header or header in {"trial_start_time", "trial_end_time", "trial_duration"}
    ]
    for row_index in range(2, trials.max_row + 1):
        raw_load = rows[row_index - 2].get("load_condition", "") if rows else ""
        row_fill = PatternFill("solid", fgColor=row_fills.get(str(raw_load), "FFFFFF" if row_index % 2 else LIGHT))
        for cell in trials[row_index]:
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10, color=TEXT, underline=None)
            cell.alignment = centered
            cell.border = Border(bottom=thin)
        for column_index in result_indexes:
            result_cell = trials.cell(row_index, column_index)
            if result_cell.value == "Correct":
                result_cell.fill = PatternFill("solid", fgColor=GREEN)
                result_cell.font = Font(name="Calibri", size=10, bold=True, color="375623", underline=None)
            elif result_cell.value == "Incorrect":
                result_cell.fill = PatternFill("solid", fgColor=RED)
                result_cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006", underline=None)
        for column_index in rt_indexes:
            cell = trials.cell(row_index, column_index)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"
        trials.row_dimensions[row_index].height = 30

    for column_index, header in enumerate(headers, start=1):
        friendly = FRIENDLY_HEADERS.get(header, header.replace("_", " ").title())
        if header == "date":
            width = 22
        elif "thought_probe_label" in header:
            width = 28
        elif "color" in header:
            width = 16
        elif "response" in header or "correct" in header:
            width = 18
        else:
            width = min(22, max(12, len(friendly) + 3))
        trials.column_dimensions[get_column_letter(column_index)].width = width
    trials.freeze_panes = "A2"
    trials.auto_filter.ref = trials.dimensions
    trials.sheet_view.zoomScale = 80
    trials.print_title_rows = "1:1"
    trials.page_setup.orientation = "landscape"
    trials.page_setup.fitToWidth = 1
    trials.sheet_properties.pageSetUpPr.fitToPage = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.stem}.temporary.xlsx")
    wb.save(temporary_path)
    temporary_path.replace(output_path)
    return output_path


def convert_data_directory(data_directory, force=False):
    data_dir = Path(data_directory).resolve()
    data_dir.mkdir(parents=True, exist_ok=True)
    converted = []
    failures = []
    for csv_path in sorted(data_dir.glob("*.csv")):
        output_path = csv_path.with_suffix(".xlsx")
        if not force and output_path.exists() and output_path.stat().st_mtime_ns >= csv_path.stat().st_mtime_ns:
            continue
        try:
            converted.append(create_professional_workbook(csv_path, output_path))
        except Exception as exc:
            failures.append((csv_path, exc))
    return converted, failures


def main():
    parser = argparse.ArgumentParser(description="Create professional Excel files from experiment CSV data.")
    parser.add_argument("--data-dir", default="data", help="Folder containing experiment CSV files")
    parser.add_argument("--force", action="store_true", help="Rebuild Excel files even when they are current")
    args = parser.parse_args()
    converted, failures = convert_data_directory(args.data_dir, force=args.force)
    for output_path in converted:
        print(f"Professional Excel saved: {output_path}")
    for csv_path, error in failures:
        print(f"Excel export failed for {csv_path.name}: {error}", file=sys.stderr)
    if not converted and not failures:
        print("No new experiment CSV files needed Excel conversion.")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
