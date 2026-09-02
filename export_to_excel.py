#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PROFESSIONAL EXCEL EXPORTER
Converts CSV data files to beautifully formatted Excel workbooks

Features:
- Color-coded headers and data
- Centered text alignment
- Auto-adjusted column widths
- Conditional formatting for accuracy
- Summary statistics sheet
- Professional styling
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime

# ============================================
# COLOR SCHEME (Modern Blue-Green Palette)
# ============================================
COLORS = {
    'header_bg': '1F4788',      # Dark blue
    'header_text': 'FFFFFF',     # White
    'participant_bg': 'E8F4F8',  # Light blue
    'trial_bg': 'FFFFFF',        # White
    'alt_row_bg': 'F8F9FA',      # Very light gray
    'correct_bg': 'D4EDDA',      # Light green
    'incorrect_bg': 'F8D7DA',    # Light red
    'warning_bg': 'FFF3CD',      # Light yellow
    'summary_bg': '28A745',      # Green
    'summary_text': 'FFFFFF',    # White
}

def create_border():
    """Create border style for cells"""
    thin_border = Side(style='thin', color='CCCCCC')
    return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

def style_header_row(ws, row=1):
    """Apply professional styling to header row"""
    for cell in ws[row]:
        cell.font = Font(bold=True, color=COLORS['header_text'], size=11)
        cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = create_border()

def style_data_rows(ws, start_row=2, end_row=None):
    """Apply styling to data rows with alternating colors"""
    if end_row is None:
        end_row = ws.max_row
    
    for idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row), start=start_row):
        # Alternating row colors
        bg_color = COLORS['alt_row_bg'] if idx % 2 == 0 else COLORS['trial_bg']
        
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.border = create_border()
            cell.font = Font(size=10)

def apply_conditional_formatting(ws, df):
    """Apply conditional formatting for accuracy columns"""
    # Find columns with 'correct' in name
    correct_cols = [col for col in df.columns if 'correct' in col.lower()]
    
    for col_name in correct_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1  # Excel is 1-indexed
            col_letter = chr(64 + col_idx)  # Convert to letter
            
            # Color coding: 1=green, 0=red, NA=yellow
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                value = str(cell.value).strip().upper()
                
                if value == '1' or value == '1.0':
                    cell.fill = PatternFill(start_color=COLORS['correct_bg'], end_color=COLORS['correct_bg'], fill_type='solid')
                    cell.font = Font(bold=True, color='155724')
                elif value == '0' or value == '0.0':
                    cell.fill = PatternFill(start_color=COLORS['incorrect_bg'], end_color=COLORS['incorrect_bg'], fill_type='solid')
                    cell.font = Font(bold=True, color='721C24')
                elif value == 'NA':
                    cell.fill = PatternFill(start_color=COLORS['warning_bg'], end_color=COLORS['warning_bg'], fill_type='solid')
                    cell.font = Font(italic=True, color='856404')

def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)  # Max width 30
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(wb, df):
    """Create a summary statistics sheet"""
    ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
    
    # Title
    ws['A1'] = 'EXPERIMENT DATA SUMMARY'
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['summary_text'])
    ws['A1'].fill = PatternFill(start_color=COLORS['summary_bg'], end_color=COLORS['summary_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    
    # Participant info
    row = 3
    ws[f'A{row}'] = 'Participant ID:'
    ws[f'B{row}'] = df['participant'].iloc[0] if 'participant' in df.columns else 'N/A'
    row += 1
    ws[f'A{row}'] = 'Experiment Date:'
    ws[f'B{row}'] = df['date'].iloc[0] if 'date' in df.columns else 'N/A'
    row += 1
    ws[f'A{row}'] = 'Age:'
    ws[f'B{row}'] = df['age'].iloc[0] if 'age' in df.columns else 'N/A'
    row += 1
    ws[f'A{row}'] = 'Gender:'
    ws[f'B{row}'] = df['gender'].iloc[0] if 'gender' in df.columns else 'N/A'
    row += 2
    
    # Statistics
    ws[f'A{row}'] = 'PERFORMANCE STATISTICS'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Headers
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'C{row}'] = 'Min'
    ws[f'D{row}'] = 'Max'
    style_header_row(ws, row)
    row += 1
    
    # Calculate statistics
    stats_row_start = row
    
    if 'memory_correct' in df.columns:
        memory_acc = df['memory_correct'].apply(lambda x: 1 if str(x).strip() == '1' else (0 if str(x).strip() == '0' else None))
        memory_acc = memory_acc.dropna()
        if len(memory_acc) > 0:
            ws[f'A{row}'] = 'Memory Accuracy'
            ws[f'B{row}'] = f"{memory_acc.mean():.2%}"
            ws[f'C{row}'] = f"{memory_acc.min():.2f}"
            ws[f'D{row}'] = f"{memory_acc.max():.2f}"
            row += 1
    
    if 'memory_rt' in df.columns:
        memory_rt = pd.to_numeric(df['memory_rt'], errors='coerce').dropna()
        if len(memory_rt) > 0:
            ws[f'A{row}'] = 'Memory RT (avg)'
            ws[f'B{row}'] = f"{memory_rt.mean():.3f}s"
            ws[f'C{row}'] = f"{memory_rt.min():.3f}s"
            ws[f'D{row}'] = f"{memory_rt.max():.3f}s"
            row += 1
    
    # Parity accuracy
    parity_cols = [col for col in df.columns if 'parity_correct' in col]
    if parity_cols:
        parity_data = []
        for col in parity_cols:
            parity_vals = df[col].apply(lambda x: 1 if str(x).strip() == '1' else (0 if str(x).strip() == '0' else None))
            parity_data.extend(parity_vals.dropna().tolist())
        if parity_data:
            parity_acc = pd.Series(parity_data)
            ws[f'A{row}'] = 'Parity Accuracy'
            ws[f'B{row}'] = f"{parity_acc.mean():.2%}"
            ws[f'C{row}'] = f"{parity_acc.min():.2f}"
            ws[f'D{row}'] = f"{parity_acc.max():.2f}"
            row += 1
    
    ws[f'A{row}'] = 'Total Trials'
    ws[f'B{row}'] = len(df)
    row += 1
    
    # Style data rows
    for r in range(stats_row_start, row):
        for cell in ws[r]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = create_border()
            if r % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS['alt_row_bg'], end_color=COLORS['alt_row_bg'], fill_type='solid')
    
    # Performance by Load
    row += 2
    ws[f'A{row}'] = 'PERFORMANCE BY LOAD CONDITION'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Load'
    ws[f'B{row}'] = 'Accuracy'
    ws[f'C{row}'] = 'RT (avg)'
    ws[f'D{row}'] = 'Trials'
    style_header_row(ws, row)
    row += 1
    
    load_row_start = row
    if 'load_condition' in df.columns and 'memory_correct' in df.columns:
        for load in sorted(df['load_condition'].unique()):
            load_data = df[df['load_condition'] == load]
            load_acc = load_data['memory_correct'].apply(lambda x: 1 if str(x).strip() == '1' else (0 if str(x).strip() == '0' else None))
            load_acc = load_acc.dropna()
            load_rt = pd.to_numeric(load_data['memory_rt'], errors='coerce').dropna()
            
            ws[f'A{row}'] = f"Load {load}"
            ws[f'B{row}'] = f"{load_acc.mean():.2%}" if len(load_acc) > 0 else 'N/A'
            ws[f'C{row}'] = f"{load_rt.mean():.3f}s" if len(load_rt) > 0 else 'N/A'
            ws[f'D{row}'] = len(load_data)
            row += 1
    
    # Style load data
    for r in range(load_row_start, row):
        for cell in ws[r]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = create_border()
            if r % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS['alt_row_bg'], end_color=COLORS['alt_row_bg'], fill_type='solid')
    
    # Auto-adjust widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def convert_csv_to_excel(csv_path):
    """Convert a single CSV file to a professionally formatted Excel file"""
    print(f"\n📊 Processing: {os.path.basename(csv_path)}")
    
    # Read CSV
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        return False
    
    # Create Excel path
    excel_path = csv_path.replace('.csv', '_FORMATTED.xlsx')
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Add summary sheet
    create_summary_sheet(wb, df)
    
    # Add data sheet
    ws_data = wb.create_sheet("Raw Data")
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
    
    # Apply styling
    style_header_row(ws_data, 1)
    style_data_rows(ws_data, 2)
    apply_conditional_formatting(ws_data, df)
    auto_adjust_column_width(ws_data)
    
    # Freeze top row
    ws_data.freeze_panes = 'A2'
    
    # Save workbook
    try:
        wb.save(excel_path)
        print(f"✅ Saved: {os.path.basename(excel_path)}")
        return True
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")
        return False

def main():
    """Main function to process all CSV files in data folder"""
    print("=" * 60)
    print("📊 PROFESSIONAL EXCEL EXPORTER")
    print("=" * 60)
    
    # Check if data folder exists
    if not os.path.exists('data'):
        print("\n❌ 'data' folder not found!")
        print("💡 Run experiments first to generate data files.")
        input("\nPress Enter to exit...")
        return
    
    # Find all CSV files
    csv_files = glob.glob('data/*_data.csv')
    
    if not csv_files:
        print("\n⚠️  No CSV data files found in 'data' folder!")
        print("💡 CSV files should end with '_data.csv'")
        input("\nPress Enter to exit...")
        return
    
    print(f"\n📁 Found {len(csv_files)} CSV file(s)")
    
    # Convert each file
    success_count = 0
    for csv_file in csv_files:
        if convert_csv_to_excel(csv_file):
            success_count += 1
    
    print("\n" + "=" * 60)
    print(f"✅ Successfully converted {success_count}/{len(csv_files)} files")
    print("=" * 60)
    print("\n💡 Formatted Excel files saved in 'data' folder")
    print("   Look for files ending with '_FORMATTED.xlsx'")
    print("\n📊 Features:")
    print("   • Color-coded headers (dark blue)")
    print("   • Alternating row colors (better readability)")
    print("   • Centered text alignment")
    print("   • Auto-adjusted column widths")
    print("   • Green = Correct | Red = Incorrect | Yellow = N/A")
    print("   • Summary statistics sheet")
    print("   • Performance by load condition")
    
    input("\n✅ Press Enter to exit...")

if __name__ == "__main__":
    main()
