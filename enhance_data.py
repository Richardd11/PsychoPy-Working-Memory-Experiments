#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Enhanced Data Formatter
Converts raw CSV data to beautiful formatted Excel file
With colors, borders, centered text, and better organization
"""

import pandas as pd
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True

def enhance_csv_to_excel(csv_file):
    """
    Convert CSV to beautifully formatted Excel file
    """
    print(f"\n{'='*70}")
    print(f"ENHANCING DATA FILE: {os.path.basename(csv_file)}")
    print(f"{'='*70}\n")
    
    # Read CSV
    print("📖 Reading CSV file...")
    df = pd.read_csv(csv_file)
    
    # Create Excel filename
    excel_file = csv_file.replace('.csv', '_FORMATTED.xlsx')
    
    # Write to Excel
    print("📝 Creating Excel file...")
    df.to_excel(excel_file, index=False, sheet_name='Data')
    
    # Load workbook for formatting
    print("🎨 Applying beautiful formatting...")
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Define colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Dark blue
    row_fill_1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Light blue
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    print("  ✓ Formatting headers...")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Get column indices for accuracy columns
    accuracy_cols = []
    for idx, col_name in enumerate(df.columns, start=1):
        if 'correct' in col_name.lower() or 'accuracy' in col_name.lower():
            accuracy_cols.append(idx)
    
    # Format data rows
    print("  ✓ Formatting data rows with alternating colors...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Alternate row colors
        fill_color = row_fill_1 if (row_idx % 2 == 0) else row_fill_2
        
        for col_idx, cell in enumerate(row, start=1):
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            
            # Color code accuracy columns
            if col_idx in accuracy_cols:
                if cell.value == 1:
                    cell.fill = correct_fill  # Green for correct
                elif cell.value == 0:
                    cell.fill = incorrect_fill  # Red for incorrect
                else:
                    cell.fill = fill_color
            else:
                cell.fill = fill_color
    
    # Auto-adjust column widths
    print("  ✓ Auto-adjusting column widths...")
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)  # Cap at 30
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    print("  ✓ Freezing header row...")
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    print("  ✓ Creating summary sheet...")
    ws_summary = wb.create_sheet("Summary")
    
    # Calculate summary statistics
    summary_data = []
    summary_data.append(['EXPERIMENT SUMMARY', ''])
    summary_data.append(['', ''])
    summary_data.append(['Participant ID:', df['participant'].iloc[0] if 'participant' in df.columns else 'N/A'])
    summary_data.append(['Total Trials:', len(df)])
    summary_data.append(['', ''])
    
    # Memory accuracy
    if 'memory_correct' in df.columns:
        memory_acc = df['memory_correct'].replace('NA', pd.NA).dropna().astype(float).mean() * 100
        summary_data.append(['Memory Accuracy:', f'{memory_acc:.2f}%'])
    
    # Parity accuracy
    parity_cols = [col for col in df.columns if 'parity_correct' in col]
    if parity_cols:
        parity_values = []
        for col in parity_cols:
            parity_values.extend(df[col].replace('NA', pd.NA).dropna().astype(float).tolist())
        if parity_values:
            parity_acc = (sum(parity_values) / len(parity_values)) * 100
            summary_data.append(['Parity Accuracy:', f'{parity_acc:.2f}%'])
    
    summary_data.append(['', ''])
    summary_data.append(['Performance by Load Condition:', ''])
    
    # By load condition
    if 'load_condition' in df.columns and 'memory_correct' in df.columns:
        for load in [0, 2, 4]:
            load_df = df[df['load_condition'] == load]
            if len(load_df) > 0:
                load_acc = load_df['memory_correct'].replace('NA', pd.NA).dropna().astype(float).mean() * 100
                summary_data.append([f'  Load {load}:', f'{load_acc:.2f}%'])
    
    summary_data.append(['', ''])
    summary_data.append(['Thought Probe Distribution:', ''])
    
    # Thought probe counts
    if 'thought_probe_label' in df.columns:
        thought_counts = df['thought_probe_label'].value_counts()
        for label, count in thought_counts.items():
            pct = (count / len(df)) * 100
            summary_data.append([f'  {label}:', f'{count} ({pct:.1f}%)'])
    
    # Write summary data
    for row_idx, row_data in enumerate(summary_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            
            # Format summary sheet
            if row_idx == 1:
                cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            elif 'Accuracy:' in str(value) or 'Load' in str(value) or row_data[0].startswith('  '):
                cell.font = Font(name='Calibri', size=10)
            else:
                cell.font = Font(name='Calibri', size=11, bold=True)
            
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
    
    # Adjust summary column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    
    # Save workbook
    print("💾 Saving formatted Excel file...")
    wb.save(excel_file)
    
    print(f"\n{'='*70}")
    print(f"✅ SUCCESS! Enhanced file created:")
    print(f"   📁 {excel_file}")
    print(f"{'='*70}\n")
    
    print("📊 Features added:")
    print("  ✓ Color-coded headers (dark blue)")
    print("  ✓ Alternating row colors (light blue/white)")
    print("  ✓ Green highlighting for correct answers")
    print("  ✓ Red highlighting for incorrect answers")
    print("  ✓ Centered text alignment")
    print("  ✓ Professional borders")
    print("  ✓ Auto-sized columns")
    print("  ✓ Frozen header row")
    print("  ✓ Summary statistics sheet")
    print()
    
    return excel_file

def main():
    """Main function to enhance all CSV files in data folder"""
    print("\n" + "="*70)
    print("CSV TO EXCEL ENHANCER")
    print("Converting raw CSV data to beautiful formatted Excel files")
    print("="*70)
    
    # Find data folder
    data_folder = Path('data')
    
    if not data_folder.exists():
        print("\n❌ Error: 'data' folder not found!")
        print("   Make sure you run this from the experiment folder.")
        input("\nPress Enter to exit...")
        return
    
    # Find all CSV files
    csv_files = list(data_folder.glob('*_detailed.csv'))
    
    if not csv_files:
        print("\n⚠️  No CSV data files found in 'data' folder.")
        print("   Run the experiment first to generate data files.")
        input("\nPress Enter to exit...")
        return
    
    print(f"\n📁 Found {len(csv_files)} CSV file(s) to enhance:\n")
    for i, file in enumerate(csv_files, 1):
        print(f"   {i}. {file.name}")
    
    print("\n" + "-"*70)
    choice = input("\nEnhance which file? (Enter number, or 'all' for all files): ").strip().lower()
    
    if choice == 'all':
        print("\n🔄 Enhancing all files...\n")
        for csv_file in csv_files:
            try:
                enhance_csv_to_excel(str(csv_file))
            except Exception as e:
                print(f"❌ Error processing {csv_file.name}: {e}")
    else:
        try:
            file_idx = int(choice) - 1
            if 0 <= file_idx < len(csv_files):
                enhance_csv_to_excel(str(csv_files[file_idx]))
            else:
                print(f"❌ Invalid choice: {choice}")
        except ValueError:
            print(f"❌ Invalid input: {choice}")
    
    print("\n" + "="*70)
    print("DONE! You can now open the Excel file in Microsoft Excel.")
    print("="*70)
    input("\nPress Enter to exit...")

if __name__ == '__main__':
    main()
